import csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

rows = []
with open('new_animation_sliders.csv', 'r') as f:
    reader = csv.DictReader(f)
    for r in reader:
        rows.append(r)

def get_family(label):
    l = label.lower().strip()
    mapping = {
        # Speed
        'animation speed': 'Speed',
        'panel speed': 'Speed',
        'lane speed': 'Speed',
        'rotation speed': 'Speed',
        'marquee speed': 'Speed',
        'breathing speed': 'Speed',
        'hover speed': 'Speed',
        'gallery speed': 'Speed',
        'scroll speed': 'Speed',
        # Border Radius
        'card border radius': 'Border Radius',
        'panel border radius': 'Border Radius',
        'image border radius': 'Border Radius',
        'tile border radius': 'Border Radius',
        'border radius': 'Border Radius',
        'page border radius': 'Border Radius',
        'item border radius': 'Border Radius',
        # Perspective
        'perspective': 'Perspective',
        # Direction
        'direction': 'Direction',
        'rotation direction': 'Direction',
        # Pointer Latency
        'pointer latency': 'Pointer Latency',
        # Spacing (gaps + padding)
        'grid gap': 'Spacing',
        'panel gap': 'Spacing',
        'card gap': 'Spacing',
        'column gap': 'Spacing',
        'gap': 'Spacing',
        'item spacing': 'Spacing',
        'stack gap': 'Spacing',
        'row gap': 'Spacing',
        'gap between items': 'Spacing',
        'panel gap (depth)': 'Spacing',
        'card spacing': 'Spacing',
        'outer padding': 'Spacing',
        'image padding': 'Spacing',
        'content padding': 'Spacing',
        'page padding': 'Spacing',
        'heading/body gap': 'Spacing',
        'image overlap': 'Spacing',
        # Dimensions (width + height)
        'card width': 'Dimensions',
        'panel width': 'Dimensions',
        'item width': 'Dimensions',
        'panel default width': 'Dimensions',
        'open width': 'Dimensions',
        'card max width': 'Dimensions',
        'content max width': 'Dimensions',
        'panel image width': 'Dimensions',
        'column width': 'Dimensions',
        'card height': 'Dimensions',
        'panel height': 'Dimensions',
        'item height': 'Dimensions',
        'panel default height': 'Dimensions',
        'open height': 'Dimensions',
        'banner height': 'Dimensions',
        'row height': 'Dimensions',
        # Size / Scale
        'ring size': 'Size / Scale',
        'tile size': 'Size / Scale',
        'card size': 'Size / Scale',
        'image size': 'Size / Scale',
        'image base size': 'Size / Scale',
        'item size': 'Size / Scale',
        'hover size': 'Size / Scale',
        'size variation': 'Size / Scale',
        'zoom level': 'Size / Scale',
        'card entry scale': 'Size / Scale',
        'mask scale': 'Size / Scale',
        '3d depth': 'Size / Scale',
        'image mask width': 'Size / Scale',
        'card border width': 'Size / Scale',
        # Typography
        'font size': 'Typography',
        'main title size': 'Typography',
        'section heading size': 'Typography',
        'text font size': 'Typography',
        'item font size': 'Typography',
        'heading font size': 'Typography',
        'featured text size': 'Typography',
        'background number size': 'Typography',
        'line height': 'Typography',
        'heading letter spacing': 'Typography',
        'stroke width': 'Typography',
        # Angle / Tilt
        'tilt angle': 'Angle / Tilt',
        'max rotation': 'Angle / Tilt',
        'max movement': 'Angle / Tilt',
        'angle': 'Angle / Tilt',
        'card rotation': 'Angle / Tilt',
        'parallax range': 'Angle / Tilt',
        # Visual Effects
        'blur intensity': 'Visual Effect',
        'parallax depth': 'Visual Effect',
        'image grayscale': 'Visual Effect',
        # Density / Count
        'density': 'Density / Count',
        'image density': 'Density / Count',
        'image count': 'Density / Count',
        # Stagger / Timing
        'pop duration': 'Stagger / Timing',
        'stagger': 'Stagger / Timing',
        'entrance stagger': 'Stagger / Timing',
        'gap / stagger': 'Stagger / Timing',
        # Spread
        'spread distance': 'Spread Distance',
        # Hover Effect
        'hover shift': 'Hover Effect',
        'hover shape': 'Hover Effect',
        # Unique niche
        'wave amplitude': 'Unique',
        'fold point': 'Unique',
        'reveal shape': 'Unique',
    }
    return mapping.get(l, label)

def short_label(label, family):
    """Return a concise cell label — strip the family name if redundant."""
    l = label.lower().strip()
    # For speed: show what kind
    speed_map = {
        'animation speed': 'Animation',
        'panel speed': 'Panel',
        'lane speed': 'Lane',
        'rotation speed': 'Rotation',
        'marquee speed': 'Marquee',
        'breathing speed': 'Breathing',
        'hover speed': 'Hover',
        'gallery speed': 'Gallery',
        'scroll speed': 'Scroll',
    }
    if l in speed_map:
        return speed_map[l]
    # For border radius: show what element
    br_map = {
        'card border radius': 'Card',
        'panel border radius': 'Panel',
        'image border radius': 'Image',
        'tile border radius': 'Tile',
        'border radius': 'Generic',
        'page border radius': 'Page',
        'item border radius': 'Item',
    }
    if l in br_map:
        return br_map[l]
    # For direction
    dir_map = {
        'direction': 'Direction',
        'rotation direction': 'Rotation',
    }
    if l in dir_map:
        return dir_map[l]
    # For spacing
    spacing_map = {
        'grid gap': 'Grid',
        'panel gap': 'Panel',
        'card gap': 'Card',
        'column gap': 'Column',
        'gap': 'Gap',
        'item spacing': 'Item',
        'stack gap': 'Stack',
        'row gap': 'Row',
        'gap between items': 'Items',
        'panel gap (depth)': 'Depth',
        'card spacing': 'Card',
        'outer padding': 'Outer Pad',
        'image padding': 'Image Pad',
        'content padding': 'Content Pad',
        'page padding': 'Page Pad',
        'heading/body gap': 'Head/Body',
        'image overlap': 'Overlap',
    }
    if l in spacing_map:
        return spacing_map[l]
    # For dimensions
    dim_map = {
        'card width': 'Card W',
        'panel width': 'Panel W',
        'item width': 'Item W',
        'panel default width': 'Panel Def W',
        'open width': 'Open W',
        'card max width': 'Card Max W',
        'content max width': 'Content Max W',
        'panel image width': 'Panel Img W',
        'column width': 'Column W',
        'card height': 'Card H',
        'panel height': 'Panel H',
        'item height': 'Item H',
        'panel default height': 'Panel Def H',
        'open height': 'Open H',
        'banner height': 'Banner H',
        'row height': 'Row H',
    }
    if l in dim_map:
        return dim_map[l]
    # Fallback: just return the original label
    return label

# Build: animation order
anim_order = []
seen_anims = set()
for r in rows:
    name = r['name']
    if name not in seen_anims:
        anim_order.append(name)
        seen_anims.add(name)

# family -> animation -> list of specific slider labels
family_anim_map = defaultdict(lambda: defaultdict(list))
for r in rows:
    fam = get_family(r['slider_label'])
    family_anim_map[fam][r['name']].append(r['slider_label'])

family_counts = {fam: len(anim_dict) for fam, anim_dict in family_anim_map.items()}
sorted_families = sorted(family_counts.keys(), key=lambda f: -family_counts[f])

# Build xlsx
wb = Workbook()
ws = wb.active
ws.title = "Slider × Animation Matrix"

GREEN = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
RED = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
HDR = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
SLD = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
CNT = PatternFill(start_color="AED6F1", end_color="AED6F1", fill_type="solid")
CHK_FONT = Font(size=8, color="1E8449")
X_FONT = Font(size=8, color="943126")
BORDER = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Header row
for c, (val, fill) in enumerate([
    ("Slider Family", HDR), ("# Used", HDR)
], start=1):
    cell = ws.cell(row=1, column=c, value=val)
    cell.font = HDR_FONT
    cell.fill = fill
    cell.alignment = LEFT if c == 1 else CENTER
    cell.border = BORDER

for ci, name in enumerate(anim_order, start=3):
    cell = ws.cell(row=1, column=ci, value=name)
    cell.font = HDR_FONT
    cell.fill = HDR
    cell.alignment = Alignment(horizontal='center', vertical='bottom',
                               text_rotation=90, wrap_text=False)
    cell.border = BORDER

# Data rows
for ri, fam in enumerate(sorted_families, start=2):
    c1 = ws.cell(row=ri, column=1, value=fam)
    c1.fill = SLD
    c1.font = Font(bold=True, size=10)
    c1.alignment = LEFT
    c1.border = BORDER

    c2 = ws.cell(row=ri, column=2, value=family_counts[fam])
    c2.fill = CNT
    c2.font = Font(bold=True, size=10)
    c2.alignment = CENTER
    c2.border = BORDER

    for ci, anim in enumerate(anim_order, start=3):
        cell = ws.cell(row=ri, column=ci)
        cell.border = BORDER
        cell.alignment = CENTER

        if anim in family_anim_map[fam]:
            labels = family_anim_map[fam][anim]
            txt = "\n".join(labels)
            cell.value = txt
            cell.fill = GREEN
            cell.font = CHK_FONT
        else:
            cell.value = "✗"
            cell.fill = RED
            cell.font = X_FONT

# Column widths — wide enough to read cell contents
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 8
for ci in range(3, 3 + len(anim_order)):
    col_letter = ws.cell(row=1, column=ci).column_letter
    ws.column_dimensions[col_letter].width = 22

ws.row_dimensions[1].height = 160
# Auto-size row heights based on max line count in that row
for ri, fam in enumerate(sorted_families, start=2):
    max_lines = 1
    for ci, anim in enumerate(anim_order, start=3):
        cell = ws.cell(row=ri, column=ci)
        if cell.value and cell.value != "✗":
            lines = cell.value.count("\n") + 1
            max_lines = max(max_lines, lines)
    ws.row_dimensions[ri].height = max(28, max_lines * 16)

ws.freeze_panes = 'C2'

########################################################################
# TAB 2 — Animation Profiles (animations as rows, families as columns)
# Adds: directory, trigger type, slider counts, family coverage %
########################################################################

ws2 = wb.create_sheet(title="Animation Profiles")

# Gather per-animation metadata from the raw CSV rows
anim_meta = {}
for r in rows:
    name = r['name']
    if name not in anim_meta:
        anim_meta[name] = {
            'dir': r['directory'],
            'triggers': r['triggers'],
            'anim_count': 0,
            'layout_count': 0,
        }
    if r['slider_tab'] == 'animation':
        anim_meta[name]['anim_count'] += 1
    else:
        anim_meta[name]['layout_count'] += 1

# Per-animation: which families it covers
anim_families = defaultdict(set)
for r in rows:
    fam = get_family(r['slider_label'])
    anim_families[r['name']].add(fam)

total_families = len(sorted_families)

# Keep original CSV order (first appearance order)
anim_sorted = anim_order

# Color palette for trigger types
TRIGGER_COLORS = {
    'scroll':  PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid"),
    'hover':   PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
    'click':   PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid"),
    'pointer': PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),
    'loop':    PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
    'mixed':   PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid"),
}

def simplify_triggers(raw):
    """Collapse raw trigger string into a human-readable interaction type."""
    t = raw.lower().replace(';', ',')
    tokens = [x.strip() for x in t.split(',')]
    has_scroll = any('viewprogress' in tk for tk in tokens)
    has_hover = any('hover' in tk for tk in tokens)
    has_click = any('click' in tk or 'toggle' in tk for tk in tokens)
    has_pointer = any('pointermove' in tk for tk in tokens)
    has_enter = any('viewenter' in tk for tk in tokens)
    has_visible = any('pagevisible' in tk for tk in tokens)

    parts = []
    if has_scroll: parts.append('Scroll')
    if has_hover: parts.append('Hover')
    if has_click: parts.append('Click')
    if has_pointer: parts.append('Pointer')
    if has_enter and not has_scroll: parts.append('Entrance')
    if has_visible and not has_scroll: parts.append('Auto-play')
    if not parts: parts.append('Other')
    return ' + '.join(parts)

def trigger_color(simple):
    s = simple.lower()
    if 'scroll' in s and '+' not in simple: return TRIGGER_COLORS['scroll']
    if 'hover' in s and '+' not in simple: return TRIGGER_COLORS['hover']
    if 'click' in s and '+' not in simple: return TRIGGER_COLORS['click']
    if 'pointer' in s and '+' not in simple: return TRIGGER_COLORS['pointer']
    if 'auto-play' in s and '+' not in simple: return TRIGGER_COLORS['loop']
    return TRIGGER_COLORS['mixed']

# Coverage bar colors (gradient from red → yellow → green)
def coverage_fill(pct):
    if pct >= 50: return PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    if pct >= 31: return PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
    return PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

# Metadata column headers
meta_headers = [
    ("Animation", 28),
    ("Folder", 24),
    ("Interaction", 18),
    ("Total", 7),
    ("Anim", 7),
    ("Layout", 7),
    ("Coverage\n(families out of 16)", 22),
    ("# Sliders", 10),
]
META_COLS = len(meta_headers)

# Write header row
for ci, (label, _) in enumerate(meta_headers, start=1):
    cell = ws2.cell(row=1, column=ci, value=label)
    cell.font = HDR_FONT
    cell.fill = HDR
    cell.alignment = CENTER
    cell.border = BORDER

for ci, fam in enumerate(sorted_families, start=META_COLS + 1):
    cell = ws2.cell(row=1, column=ci, value=fam)
    cell.font = HDR_FONT
    cell.fill = HDR
    cell.alignment = CENTER
    cell.border = BORDER

# Data rows
for ri, anim in enumerate(anim_sorted, start=2):
    meta = anim_meta[anim]
    total = meta['anim_count'] + meta['layout_count']
    cov_count = len(anim_families[anim])
    cov_pct = round(100 * cov_count / total_families)
    trig_simple = simplify_triggers(meta['triggers'])

    # Column A: Animation name
    c = ws2.cell(row=ri, column=1, value=anim)
    c.font = Font(bold=True, size=10)
    c.fill = SLD
    c.alignment = LEFT
    c.border = BORDER

    # Column B: Directory (verbatim)
    c = ws2.cell(row=ri, column=2, value=meta['dir'])
    c.font = Font(size=9, italic=True)
    c.alignment = CENTER
    c.border = BORDER

    # Column C: Interaction type
    c = ws2.cell(row=ri, column=3, value=trig_simple)
    c.font = Font(size=9, bold=True)
    c.fill = trigger_color(trig_simple)
    c.alignment = CENTER
    c.border = BORDER

    # Column D: Total sliders
    c = ws2.cell(row=ri, column=4, value=total)
    c.font = Font(size=10, bold=True)
    c.alignment = CENTER
    c.border = BORDER

    # Column E: Animation count
    c = ws2.cell(row=ri, column=5, value=meta['anim_count'])
    c.font = Font(size=9, color="2E86C1")
    c.alignment = CENTER
    c.border = BORDER

    # Column F: Layout count
    c = ws2.cell(row=ri, column=6, value=meta['layout_count'])
    c.font = Font(size=9, color="7D3C98")
    c.alignment = CENTER
    c.border = BORDER

    # Column G: Coverage
    c = ws2.cell(row=ri, column=7, value=f"{cov_count} / {total_families}  ({cov_pct}%)")
    c.font = Font(size=9, bold=True)
    c.fill = coverage_fill(cov_pct)
    c.alignment = CENTER
    c.border = BORDER

    # Column H: # Sliders (actual individual slider count)
    c = ws2.cell(row=ri, column=8, value=total)
    c.font = Font(size=10, bold=True)
    c.fill = CNT
    c.alignment = CENTER
    c.border = BORDER

    # Family columns
    for ci, fam in enumerate(sorted_families, start=META_COLS + 1):
        cell = ws2.cell(row=ri, column=ci)
        cell.border = BORDER
        cell.alignment = CENTER

        if anim in family_anim_map[fam]:
            labels = family_anim_map[fam][anim]
            cell.value = "\n".join(labels)
            cell.fill = GREEN
            cell.font = CHK_FONT
        else:
            cell.value = "✗"
            cell.fill = RED
            cell.font = X_FONT

# Column widths
for ci, (_, w) in enumerate(meta_headers, start=1):
    ws2.column_dimensions[ws2.cell(row=1, column=ci).column_letter].width = w

for ci in range(META_COLS + 1, META_COLS + 1 + len(sorted_families)):
    ws2.column_dimensions[ws2.cell(row=1, column=ci).column_letter].width = 22

# Row heights
ws2.row_dimensions[1].height = 30
for ri, anim in enumerate(anim_sorted, start=2):
    max_lines = 1
    for ci in range(META_COLS + 1, META_COLS + 1 + len(sorted_families)):
        val = ws2.cell(row=ri, column=ci).value
        if val and val != "✗":
            max_lines = max(max_lines, val.count("\n") + 1)
    ws2.row_dimensions[ri].height = max(24, max_lines * 16)

# Freeze: keep header row + first meta columns visible
ws2.freeze_panes = ws2.cell(row=2, column=META_COLS + 1).coordinate

########################################################################
wb.save('slider_animation_matrix.xlsx')
print(f"Done: slider_animation_matrix.xlsx")
print(f"  Tab 1: {len(sorted_families)} slider families × {len(anim_order)} animations")
print(f"  Tab 2: {len(anim_sorted)} animations × {len(sorted_families)} slider families + metadata")
for fam in sorted_families:
    print(f"  {family_counts[fam]:>2}  {fam}")
